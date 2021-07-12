import openpyxl

wb = openpyxl.Workbook()

sheet2 = wb.create_sheet()

sheet2.title = 'WorkOrders_June2021'

wb.create_sheet(index=0, title='Shipped_WorkOrders')

wb.save('WorkOrders_Clonals.csv')
