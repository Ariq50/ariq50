import openpyxl
import os

os.chdir('c:\\user\\sophytes\\documents')

workbook = openpyxl.load_workbook('HitPickFile.csv')
workbook.get_sheet_names()

cell = sheet[HitPickFile]

cell.value

sheet.cell(row=1, column=2)

for i in range(2, 31):
    print(i, sheet.cell(row=i, column=2).value
    
